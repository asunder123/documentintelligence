
# ingestion/extractor.py
# Document-type aware text extraction
# Responsibility: convert files → clean text (ONLY)

import json
import re
import zipfile
from typing import Any

# Optional libraries (graceful degradation)
try:
    import pdfplumber
except Exception:
    pdfplumber = None

try:
    import openpyxl
    from openpyxl.utils.cell import range_boundaries
except Exception:
    openpyxl = None


# ============================================================
# Utilities
# ============================================================

def _safe_decode(data: bytes) -> str:
    try:
        return data.decode("utf-8")
    except Exception:
        return data.decode("latin1", errors="ignore")


def _normalize_whitespace(text: str) -> str:
    """Normalize whitespace across the whole text (use sparingly)."""
    return re.sub(r"\s+", " ", text).strip()


def _normalize_line(s: str) -> str:
    """Normalize a single line: keep newlines intact across the whole document."""
    return re.sub(r"[ \t]+", " ", s).strip()


# ============================================================
# Plain text formats
# ============================================================

def _extract_text_plain(uploaded_file) -> str:
    return _safe_decode(uploaded_file.read())


# ============================================================
# JSON
# ============================================================

def _extract_text_json(uploaded_file) -> str:
    raw = _safe_decode(uploaded_file.read())
    try:
        obj = json.loads(raw)
        return json.dumps(obj, indent=2)
    except Exception:
        return raw


# ============================================================
# HTML
# ============================================================

def _extract_text_html(uploaded_file) -> str:
    text = _safe_decode(uploaded_file.read())

    # Remove scripts and styles
    text = re.sub(r"<script.*?>.*?</script>", " ", text, flags=re.S | re.I)
    text = re.sub(r"<style.*?>.*?</style>", " ", text, flags=re.S | re.I)

    # Strip tags
    text = re.sub(r"<[^>]+>", " ", text)

    # Keep as a single paragraph; HTML often has visual-only line breaks
    return _normalize_whitespace(text)


# ============================================================
# DOCX
# ============================================================

def _extract_text_docx(uploaded_file) -> str:
    try:
        with zipfile.ZipFile(uploaded_file) as z:
            xml = z.read("word/document.xml").decode("utf-8", errors="ignore")
            text = re.sub(r"<[^>]+>", " ", xml)
            return _normalize_whitespace(text)
    except Exception:
        return ""


# ============================================================
# PDF
# ============================================================

def _extract_text_pdf(uploaded_file) -> str:
    if pdfplumber is not None:
        try:
            with pdfplumber.open(uploaded_file) as pdf:
                pages = []
                for page in pdf.pages:
                    t = page.extract_text()
                    if t:
                        pages.append(t)
                # Preserve page breaks as newlines
                return "\n\n".join(_normalize_line(p) for p in pages if p)
        except Exception:
            pass

    # Fallback: raw decode (lossy but safe)
    return _safe_decode(uploaded_file.read())


# ============================================================
# XLSX (Excel) — Preserve table schemas where available
# ============================================================

def _extract_text_xlsx(uploaded_file) -> str:
    """
    Excel extractor that preserves schema:
      - Detects native Excel tables (ListObjects) and emits `Schema` + per-row records
      - If no tables, uses header detection and still emits structured records
      - Prefers normal mode (merged ranges/table metadata), falls back to read-only
      - Preserves newlines by normalizing per line; handles hyperlinks/booleans/numbers
    """
    if openpyxl is None:
        return ""

    # Tunables
    MAX_SHEETS = 50
    MAX_ROWS_PER_SHEET = 50000
    MAX_COLS_PER_SHEET = 512
    SKIP_HIDDEN_SHEETS = True
    MIN_STRUCTURED_RECORDS = 3  # below this, also emit simple rows for completeness

    def format_cell(cell) -> str:
        v = getattr(cell, "value", None)
        if v is None:
            return ""
        target = None
        try:
            if getattr(cell, "hyperlink", None):
                target = getattr(cell.hyperlink, "target", None)
        except Exception:
            target = None

        if isinstance(v, bool):
            s = "true" if v else "false"
        elif isinstance(v, (int, float)):
            # human-friendly no-scientific notation
            s = f"{v:.15g}"
        else:
            s = str(v)
        s = s.strip()

        if target and target not in s:
            s = f"{s} ({target})"
        return s

    def row_is_empty(cells) -> bool:
        return all((format_cell(c) == "") for c in cells)

    def detect_header_row(sheet) -> int | None:
        """
        Heuristic for non-table sheets:
          - First non-empty row with >=2 non-empty cells
          - Either sufficiently textual OR many short tokens (IDs/codes)
        """
        non_empty_rows = 0
        max_row = min(sheet.max_row, MAX_ROWS_PER_SHEET)
        max_col = min(sheet.max_column, MAX_COLS_PER_SHEET)
        for r in sheet.iter_rows(max_row=max_row, max_col=max_col):
            if row_is_empty(r):
                continue
            non_empty_rows += 1
            values = [format_cell(c) for c in r]
            non_empty = [v for v in values if v != ""]
            if len(non_empty) < 2:
                continue
            textish = sum(1 for v in non_empty if not re.fullmatch(r"[+-]?\d+(?:\.\d+)?", v))
            shortish = sum(1 for v in non_empty if len(v) <= 8)
            ratio_text = textish / max(1, len(non_empty))
            ratio_short = shortish / max(1, len(non_empty))
            if (ratio_text >= 0.45) or (ratio_short >= 0.60):
                return r[0].row
            if non_empty_rows > 150:
                break
        return None

    def get_tables(sheet):
        """
        Return a list of table-like objects with .name and .ref.
        Works across openpyxl versions (ws.tables dict vs ws._tables list).
        """
        tables = []
        try:
            # openpyxl >= 3.1: sheet.tables is a dict {name: Table}
            tbls = getattr(sheet, "tables", None)
            if isinstance(tbls, dict):
                tables = list(tbls.values())
        except Exception:
            pass
        try:
            # older: sheet._tables is a list [Table,...]
            if not tables:
                tables = getattr(sheet, "_tables", []) or []
        except Exception:
            pass
        # Filter to items that have both name and ref
        return [t for t in tables if hasattr(t, "name") and hasattr(t, "ref")]

    def read_table_block(sheet, table_obj):
        """
        Read a native Excel table by its ref; return (headers, rows_as_lists).
        """
        try:
            min_col, min_row, max_col, max_row = range_boundaries(table_obj.ref)
        except Exception:
            # Fallback: try to parse ref manually like "A1:D100"
            try:
                from openpyxl.utils import coordinate_from_string, column_index_from_string
                start, end = table_obj.ref.split(":")
                sc, sr = coordinate_from_string(start)
                ec, er = coordinate_from_string(end)
                min_col = column_index_from_string(sc)
                min_row = sr
                max_col = column_index_from_string(ec)
                max_row = er
            except Exception:
                return [], []

        # Header row is min_row in Excel tables
        headers = []
        for c in range(min_col, max_col + 1):
            cell = sheet.cell(row=min_row, column=c)
            hdr = format_cell(cell)
            headers.append(hdr if hdr else f"Column{c}")

        # Deduplicate headers by suffixing duplicates
        seen = {}
        dedup_headers = []
        for h in headers:
            if h not in seen:
                dedup_headers.append(h)
                seen[h] = 1
            else:
                seen[h] += 1
                dedup_headers.append(f"{h}::{seen[h]}")
        headers = dedup_headers

        # Data rows
        rows = []
        for r in range(min_row + 1, max_row + 1):
            row_vals = []
            empty = True
            for c in range(min_col, max_col + 1):
                v = format_cell(sheet.cell(row=r, column=c))
                if v:
                    empty = False
                row_vals.append(v)
            if not empty:
                rows.append(row_vals)
        return headers, rows

    # Try normal mode first to retain merged cell and table metadata
    wb = None
    try:
        wb = openpyxl.load_workbook(uploaded_file, data_only=True, read_only=False)
    except Exception:
        try:
            # Fallback: read-only (fast; tables may still be accessible in newer versions)
            uploaded_file.seek(0)
            wb = openpyxl.load_workbook(uploaded_file, data_only=True, read_only=True)
        except Exception:
            return ""

    out_lines: list[str] = []
    sheet_count = 0

    for sheet in wb.worksheets:
        sheet_count += 1
        if sheet_count > MAX_SHEETS:
            break

        if SKIP_HIDDEN_SHEETS and getattr(sheet, "sheet_state", "") == "hidden":
            continue

        out_lines.append(_normalize_line(f"Sheet: {sheet.title}"))

        tables = get_tables(sheet)

        if tables:
            # Emit each table with schema intact
            for t in tables:
                headers, rows = read_table_block(sheet, t)
                out_lines.append(_normalize_line(f"Table: {t.name}"))
                if headers:
                    out_lines.append(_normalize_line("Schema: " + " | ".join(headers)))
                record_id = 0
                for row_vals in rows:
                    record_id += 1
                    out_lines.append(_normalize_line(f"Record {record_id}:"))
                    for idx, val in enumerate(row_vals):
                        if not val:
                            continue
                        key = headers[idx] if idx < len(headers) else f"Column{idx+1}"
                        out_lines.append(_normalize_line(f"- {key}: {val}"))
                out_lines.append("")  # blank line after table
        else:
            # No native tables; use header detection to keep schema
            header_row_idx = detect_header_row(sheet)
            max_row = min(sheet.max_row, MAX_ROWS_PER_SHEET)
            max_col = min(sheet.max_column, MAX_COLS_PER_SHEET)

            if header_row_idx is not None:
                headers = []
                for c in range(1, max_col + 1):
                    hdr = _normalize_line(format_cell(sheet.cell(row=header_row_idx, column=c)))
                    headers.append(hdr if hdr else f"Column{c}")
                # Deduplicate headers by suffixing duplicates
                seen = {}
                dedup_headers = []
                for h in headers:
                    if h not in seen:
                        dedup_headers.append(h)
                        seen[h] = 1
                    else:
                        seen[h] += 1
                        dedup_headers.append(f"{h}::{seen[h]}")
                headers = dedup_headers

                out_lines.append(_normalize_line("Schema: " + " | ".join(headers)))

                record_id = 0
                structured_records = 0
                for r in range(header_row_idx + 1, max_row + 1):
                    row_vals = [format_cell(sheet.cell(row=r, column=c)) for c in range(1, max_col + 1)]
                    non_empty = [v for v in row_vals if v]
                    if not non_empty:
                        continue
                    record_id += 1
                    structured_records += 1
                    out_lines.append(_normalize_line(f"Record {record_id}:"))
                    for idx, val in enumerate(row_vals):
                        if not val:
                            continue
                        key = headers[idx] if idx < len(headers) else f"Column{idx+1}"
                        out_lines.append(_normalize_line(f"- {key}: {val}"))

                # If too few structured records, also emit simple rows for completeness
                if structured_records < MIN_STRUCTURED_RECORDS:
                    out_lines.append(_normalize_line("Rows (fallback):"))
                    for r in range(1, max_row + 1):
                        row_vals = [format_cell(sheet.cell(row=r, column=c)) for c in range(1, max_col + 1)]
                        non_empty = [v for v in row_vals if v]
                        if non_empty:
                            out_lines.append(_normalize_line(" | ".join(non_empty)))
            else:
                # Last resort: emit readable rows without losing content
                for r in range(1, min(sheet.max_row, MAX_ROWS_PER_SHEET) + 1):
                    row_vals = [format_cell(sheet.cell(row=r, column=c)) for c in range(1, min(sheet.max_column, MAX_COLS_PER_SHEET) + 1)]
                    non_empty = [v for v in row_vals if v]
                    if non_empty:
                        out_lines.append(_normalize_line(" | ".join(non_empty)))

        out_lines.append("")  # blank line between sheets

    # Preserve line breaks; normalize per line only
    return "\n".join(out_lines)


# ============================================================
# Public API
# ============================================================

def extract_text_from_file(uploaded_file: Any) -> str:
    """
    Master extractor.
    Routes to document-type specific logic.
    Always returns a string.
    """

    if uploaded_file is None:
        return ""

    name = uploaded_file.name.lower()

    if name.endswith((".txt", ".md", ".csv")):
        return _extract_text_plain(uploaded_file)

    if name.endswith(".json"):
        return _extract_text_json(uploaded_file)

    if name.endswith((".html", ".htm")):
        return _extract_text_html(uploaded_file)

    if name.endswith(".docx"):
        return _extract_text_docx(uploaded_file)

    if name.endswith(".pdf"):
        return _extract_text_pdf(uploaded_file)

    if name.endswith(".xlsx"):
        return _extract_text_xlsx(uploaded_file)

    # Fallback for unknown formats
    return _safe_decode(uploaded_file.read())
